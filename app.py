from flask import Flask, render_template, request, redirect, session, flash, send_file
from werkzeug.security import generate_password_hash, check_password_hash
from datetime import datetime, timedelta
from decimal import Decimal
from collections import OrderedDict
from models import conectar, criar_tabelas
import io
import smtplib
import random
import string
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart

app = Flask(__name__)
app.secret_key = "super_secret_key"

app.config["SESSION_PERMANENT"] = True
app.permanent_session_lifetime = timedelta(hours=8)

EMAIL_REMETENTE = "SEUEMAIL@gmail.com"
SENHA_APP = "SUA_SENHA_DE_APP"

criar_tabelas()


# ==================================================
# FUNÇÕES BASE
# ==================================================
def login_obrigatorio():
    return "user" in session


def cargo_usuario():
    return session.get("cargo", "Funcionario")


def exigir_cargo(*cargos):
    return cargo_usuario() in cargos


def gerar_senha_temporaria(tamanho=8):
    caracteres = string.ascii_letters + string.digits
    return "".join(random.choice(caracteres) for _ in range(tamanho))


def limpar_cpf(cpf):
    return "".join(filter(str.isdigit, cpf or ""))


def formatar_cpf(cpf):
    cpf = limpar_cpf(cpf)
    if len(cpf) != 11:
        return cpf
    return f"{cpf[:3]}.{cpf[3:6]}.{cpf[6:9]}-{cpf[9:]}"


def validar_cpf(cpf):
    cpf = limpar_cpf(cpf)

    if len(cpf) != 11 or cpf == cpf[0] * 11:
        return False

    soma = sum(int(cpf[i]) * (10 - i) for i in range(9))
    dig1 = (soma * 10 % 11) % 10

    soma = sum(int(cpf[i]) * (11 - i) for i in range(10))
    dig2 = (soma * 10 % 11) % 10

    return cpf[-2:] == f"{dig1}{dig2}"


def validar_forca_senha(senha):
    if not senha or len(senha) < 6:
        return False, "A senha deve ter pelo menos 6 caracteres."

    tem_letra = any(c.isalpha() for c in senha)
    tem_numero = any(c.isdigit() for c in senha)

    if not tem_letra or not tem_numero:
        return False, "A senha deve ter letras e números."

    return True, ""


# ==================================================
# ADMIN PADRÃO
# ==================================================
def criar_admin_padrao():
    con = conectar()
    cur = con.cursor()

    # garante coluna email em usuarios
    cur.execute("""
        ALTER TABLE usuarios
        ADD COLUMN IF NOT EXISTS email VARCHAR(150)
    """)

    cur.execute("SELECT id, cargo FROM usuarios WHERE nome=%s", ("Maycon",))
    existe = cur.fetchone()

    senha_hash = generate_password_hash("12123$")

    if not existe:
        cur.execute("""
            INSERT INTO usuarios (nome, senha, cargo, email)
            VALUES (%s, %s, %s, %s)
        """, ("Maycon", senha_hash, "Admin", "SEUEMAIL@gmail.com"))
    else:
        cur.execute("""
            UPDATE usuarios
            SET cargo=%s
            WHERE nome=%s
        """, ("Admin", "Maycon"))

    con.commit()
    cur.close()
    con.close()


criar_admin_padrao()


# ==================================================
# MÉTRICAS
# ==================================================
def calcular_metricas():
    con = conectar()
    cur = con.cursor()

    cur.execute("SELECT COUNT(*) FROM clientes")
    total_clientes = cur.fetchone()[0]

    cur.execute("SELECT COUNT(*) FROM clientes WHERE status='Ativo'")
    clientes_ativos = cur.fetchone()[0]

    cur.execute("SELECT COALESCE(SUM(valor),0) FROM financeiro WHERE tipo='Entrada'")
    total_entradas = cur.fetchone()[0] or 0

    cur.execute("SELECT COALESCE(SUM(valor),0) FROM financeiro WHERE tipo='Saída'")
    total_saidas = cur.fetchone()[0] or 0

    saldo = Decimal(total_entradas) - Decimal(total_saidas)

    cur.execute("""
        SELECT data_lancamento, tipo, valor
        FROM financeiro
        ORDER BY data_lancamento ASC
    """)
    dados = cur.fetchall()

    cur.close()
    con.close()

    meses = OrderedDict()
    hoje = datetime.today()

    for i in range(5, -1, -1):
        ref = hoje.replace(day=1) - timedelta(days=30 * i)
        chave = ref.strftime("%m/%Y")
        meses[chave] = Decimal("0")

    for data_lancamento, tipo, valor in dados:
        chave = data_lancamento.strftime("%m/%Y")
        if chave in meses:
            if tipo == "Entrada":
                meses[chave] += Decimal(valor)
            else:
                meses[chave] -= Decimal(valor)

    return {
        "total_clientes": total_clientes,
        "clientes_ativos": clientes_ativos,
        "total_entradas": float(total_entradas),
        "total_saidas": float(total_saidas),
        "saldo": float(saldo),
        "grafico_labels": list(meses.keys()),
        "grafico_valores": [float(v) for v in meses.values()]
    }


# ==================================================
# LOGIN
# ==================================================
@app.route("/", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        nome = (request.form.get("nome") or "").strip()
        senha = request.form.get("senha") or ""

        con = conectar()
        cur = con.cursor()

        cur.execute("""
            SELECT id, nome, senha, cargo
            FROM usuarios
            WHERE nome=%s
        """, (nome,))
        usuario = cur.fetchone()

        cur.close()
        con.close()

        if usuario and check_password_hash(usuario[2], senha):
            session["user_id"] = usuario[0]
            session["user"] = usuario[1]
            session["cargo"] = usuario[3] if usuario[3] else "Funcionario"
            return redirect("/dashboard")

        flash("Login inválido.")
        return redirect("/")

    return render_template("login.html")


# ==================================================
# RECUPERAR SENHA
# ==================================================
@app.route("/recuperar-senha", methods=["GET", "POST"])
def recuperar_senha():
    if request.method == "POST":
        nome = (request.form.get("nome") or "").strip()
        email = (request.form.get("email") or "").strip()

        if not nome or not email:
            flash("Preencha nome e e-mail.")
            return redirect("/recuperar-senha")

        con = conectar()
        cur = con.cursor()

        cur.execute("""
            ALTER TABLE usuarios
            ADD COLUMN IF NOT EXISTS email VARCHAR(150)
        """)

        cur.execute("""
            SELECT id, nome, email
            FROM usuarios
            WHERE nome=%s AND email=%s
        """, (nome, email))
        usuario = cur.fetchone()

        if not usuario:
            con.commit()
            cur.close()
            con.close()
            flash("Usuário ou e-mail não encontrado.")
            return redirect("/recuperar-senha")

        nova_senha = gerar_senha_temporaria(8)
        senha_hash = generate_password_hash(nova_senha)

        cur.execute("""
            UPDATE usuarios
            SET senha=%s
            WHERE id=%s
        """, (senha_hash, usuario[0]))

        con.commit()
        cur.close()
        con.close()

        try:
            msg = MIMEMultipart()
            msg["From"] = EMAIL_REMETENTE
            msg["To"] = email
            msg["Subject"] = "Recuperação de senha - Empresa+"

            texto = f"""
Olá, {nome}!

Sua nova senha temporária é:

{nova_senha}

Entre no sistema e depois altere sua senha.

Atenciosamente,
Equipe Empresa+
"""
            msg.attach(MIMEText(texto, "plain"))

            servidor = smtplib.SMTP("smtp.gmail.com", 587)
            servidor.starttls()
            servidor.login(EMAIL_REMETENTE, SENHA_APP)
            servidor.sendmail(EMAIL_REMETENTE, email, msg.as_string())
            servidor.quit()

            flash("Nova senha enviada para o seu e-mail.")
        except Exception as e:
            flash(f"Senha resetada, mas houve erro ao enviar e-mail: {str(e)}")

        return redirect("/")

    return render_template("recuperar_senha.html")


# ==================================================
# DASHBOARD
# ==================================================
@app.route("/dashboard")
def dashboard():
    if not login_obrigatorio():
        return redirect("/")

    busca = request.args.get("busca", "")

    con = conectar()
    cur = con.cursor()

    if busca:
        cur.execute("""
            SELECT id, nome, cpf, nascimento, telefone, email, status, observacoes
            FROM clientes
            WHERE nome ILIKE %s OR email ILIKE %s
            ORDER BY id DESC
        """, (f"%{busca}%", f"%{busca}%"))
    else:
        cur.execute("""
            SELECT id, nome, cpf, nascimento, telefone, email, status, observacoes
            FROM clientes
            ORDER BY id DESC
        """)
    clientes = cur.fetchall()

    cur.execute("""
        SELECT id, descricao, tipo, valor, data_lancamento
        FROM financeiro
        ORDER BY id DESC
        LIMIT 20
    """)
    financeiro = cur.fetchall()

    cur.execute("""
        ALTER TABLE usuarios
        ADD COLUMN IF NOT EXISTS email VARCHAR(150)
    """)

    cur.execute("""
        SELECT id, nome, cargo, email
        FROM usuarios
        ORDER BY id DESC
    """)
    usuarios_sistema = cur.fetchall()

    con.commit()
    cur.close()
    con.close()

    metricas = calcular_metricas()

    return render_template(
        "dashboard.html",
        usuario_logado=session["user"],
        cargo=session["cargo"],
        busca=busca,
        clientes=clientes,
        financeiro=financeiro,
        usuarios_sistema=usuarios_sistema,
        **metricas
    )


# ==================================================
# USUÁRIOS DO SISTEMA
# ==================================================
@app.route("/usuarios/cadastrar", methods=["POST"])
def cadastrar_usuario():
    if not login_obrigatorio():
        return redirect("/")

    if not exigir_cargo("Admin"):
        flash("Somente Admin pode criar usuários do sistema.")
        return redirect("/dashboard")

    nome = (request.form.get("nome") or "").strip()
    email = (request.form.get("email") or "").strip()
    senha = request.form.get("senha") or ""
    cargo = (request.form.get("cargo") or "").strip()

    if not nome or not email or not senha or not cargo:
        flash("Preencha nome, e-mail, senha e cargo.")
        return redirect("/dashboard")

    if cargo not in ["Admin", "Gerente", "Funcionario"]:
        flash("Cargo inválido.")
        return redirect("/dashboard")

    ok, msg = validar_forca_senha(senha)
    if not ok:
        flash(msg)
        return redirect("/dashboard")

    con = conectar()
    cur = con.cursor()

    cur.execute("""
        ALTER TABLE usuarios
        ADD COLUMN IF NOT EXISTS email VARCHAR(150)
    """)

    cur.execute("SELECT id FROM usuarios WHERE nome=%s", (nome,))
    existe = cur.fetchone()

    if existe:
        con.commit()
        cur.close()
        con.close()
        flash("Já existe usuário com esse nome.")
        return redirect("/dashboard")

    senha_hash = generate_password_hash(senha)

    cur.execute("""
        INSERT INTO usuarios (nome, senha, cargo, email)
        VALUES (%s, %s, %s, %s)
    """, (nome, senha_hash, cargo, email))

    con.commit()
    cur.close()
    con.close()

    flash(f"Usuário {cargo} criado com sucesso.")
    return redirect("/dashboard")


@app.route("/usuarios/excluir/<int:id>")
def excluir_usuario(id):
    if not login_obrigatorio():
        return redirect("/")

    if not exigir_cargo("Admin"):
        flash("Somente Admin pode excluir usuários.")
        return redirect("/dashboard")

    if id == session.get("user_id"):
        flash("Você não pode excluir seu próprio usuário.")
        return redirect("/dashboard")

    con = conectar()
    cur = con.cursor()

    cur.execute("DELETE FROM usuarios WHERE id=%s", (id,))
    con.commit()

    cur.close()
    con.close()

    flash("Usuário excluído.")
    return redirect("/dashboard")


# ==================================================
# CLIENTES
# ==================================================
@app.route("/clientes/cadastrar", methods=["POST"])
def cadastrar_cliente():
    if not login_obrigatorio():
        return redirect("/")

    nome = request.form.get("nome")
    cpf = request.form.get("cpf")
    nascimento = request.form.get("nascimento")
    telefone = request.form.get("telefone")
    email = request.form.get("email")
    status = request.form.get("status")
    observacoes = request.form.get("observacoes")
    senha_cliente = request.form.get("senha_cliente")

    if not validar_cpf(cpf):
        flash("CPF inválido.")
        return redirect("/dashboard")

    ok, msg = validar_forca_senha(senha_cliente)
    if not ok:
        flash(msg)
        return redirect("/dashboard")

    senha_hash = generate_password_hash(senha_cliente)

    con = conectar()
    cur = con.cursor()

    cur.execute("""
        INSERT INTO clientes
        (nome, cpf, nascimento, telefone, email, status, observacoes, senha_cliente)
        VALUES (%s,%s,%s,%s,%s,%s,%s,%s)
    """, (
        nome,
        formatar_cpf(cpf),
        nascimento,
        telefone,
        email,
        status,
        observacoes,
        senha_hash
    ))

    con.commit()
    cur.close()
    con.close()

    flash("Cliente cadastrado.")
    return redirect("/dashboard")


@app.route("/clientes/editar/<int:id>", methods=["GET", "POST"])
def editar_cliente(id):
    if not login_obrigatorio():
        return redirect("/")

    con = conectar()
    cur = con.cursor()

    if request.method == "POST":
        nome = (request.form.get("nome") or "").strip()
        cpf = request.form.get("cpf") or ""
        nascimento = request.form.get("nascimento") or ""
        telefone = (request.form.get("telefone") or "").strip()
        email = (request.form.get("email") or "").strip()
        status = (request.form.get("status") or "").strip()
        observacoes = (request.form.get("observacoes") or "").strip()
        nova_senha_cliente = request.form.get("senha_cliente") or ""

        if not nome or not cpf or not nascimento or not telefone or not email or not status:
            flash("Preencha os campos obrigatórios.")
            cur.close()
            con.close()
            return redirect(f"/clientes/editar/{id}")

        if not validar_cpf(cpf):
            flash("CPF inválido.")
            cur.close()
            con.close()
            return redirect(f"/clientes/editar/{id}")

        cpf_formatado = formatar_cpf(cpf)

        if nova_senha_cliente:
            senha_ok, mensagem_senha = validar_forca_senha(nova_senha_cliente)
            if not senha_ok:
                flash(f"Nova senha inválida. {mensagem_senha}")
                cur.close()
                con.close()
                return redirect(f"/clientes/editar/{id}")

            senha_hash = generate_password_hash(nova_senha_cliente)
            cur.execute("""
                UPDATE clientes
                SET nome=%s, cpf=%s, nascimento=%s, telefone=%s, email=%s,
                    status=%s, observacoes=%s, senha_cliente=%s
                WHERE id=%s
            """, (
                nome, cpf_formatado, nascimento, telefone, email,
                status, observacoes, senha_hash, id
            ))
        else:
            cur.execute("""
                UPDATE clientes
                SET nome=%s, cpf=%s, nascimento=%s, telefone=%s, email=%s,
                    status=%s, observacoes=%s
                WHERE id=%s
            """, (
                nome, cpf_formatado, nascimento, telefone, email,
                status, observacoes, id
            ))

        con.commit()
        cur.close()
        con.close()

        flash("Cliente atualizado com sucesso.")
        return redirect("/dashboard")

    cur.execute("""
        SELECT id, nome, cpf, nascimento, telefone, email, status, observacoes
        FROM clientes
        WHERE id=%s
    """, (id,))
    cliente = cur.fetchone()

    cur.close()
    con.close()

    if not cliente:
        flash("Cliente não encontrado.")
        return redirect("/dashboard")

    return render_template("editar_cliente.html", cliente=cliente)


@app.route("/clientes/excluir/<int:id>")
def excluir_cliente(id):
    if not login_obrigatorio():
        return redirect("/")

    if not exigir_cargo("Admin", "Gerente"):
        flash("Sem permissão.")
        return redirect("/dashboard")

    con = conectar()
    cur = con.cursor()

    cur.execute("DELETE FROM clientes WHERE id=%s", (id,))
    con.commit()

    cur.close()
    con.close()

    flash("Cliente excluído.")
    return redirect("/dashboard")


# ==================================================
# FINANCEIRO
# ==================================================
@app.route("/financeiro/lancar", methods=["POST"])
def lancar_financeiro():
    if not login_obrigatorio():
        return redirect("/")

    if not exigir_cargo("Admin", "Gerente"):
        flash("Sem permissão.")
        return redirect("/dashboard")

    descricao = request.form.get("descricao")
    tipo = request.form.get("tipo")
    valor = request.form.get("valor").replace(",", ".")
    data_lancamento = request.form.get("data_lancamento")

    try:
        valor = Decimal(valor)
    except:
        flash("Valor inválido.")
        return redirect("/dashboard")

    con = conectar()
    cur = con.cursor()

    cur.execute("""
        INSERT INTO financeiro (descricao, tipo, valor, data_lancamento)
        VALUES (%s,%s,%s,%s)
    """, (descricao, tipo, valor, data_lancamento))

    con.commit()
    cur.close()
    con.close()

    flash("Lançamento registrado.")
    return redirect("/dashboard")


# ==================================================
# E-MAIL CLIENTE
# ==================================================
@app.route("/enviar/email/<int:id>")
def enviar_email_cliente(id):
    if not login_obrigatorio():
        return redirect("/")

    con = conectar()
    cur = con.cursor()

    cur.execute("""
        SELECT nome, email
        FROM clientes
        WHERE id=%s
    """, (id,))
    cliente = cur.fetchone()

    cur.close()
    con.close()

    if not cliente:
        flash("Cliente não encontrado.")
        return redirect("/dashboard")

    nome = cliente[0]
    email = cliente[1]

    try:
        msg = MIMEMultipart()
        msg["From"] = EMAIL_REMETENTE
        msg["To"] = email
        msg["Subject"] = "Contato da Empresa+"

        texto = f"""
Olá, {nome}!

Estamos entrando em contato pela Empresa+.

Atenciosamente,
Equipe Empresa+
"""

        msg.attach(MIMEText(texto, "plain"))

        servidor = smtplib.SMTP("smtp.gmail.com", 587)
        servidor.starttls()
        servidor.login(EMAIL_REMETENTE, SENHA_APP)
        servidor.sendmail(EMAIL_REMETENTE, email, msg.as_string())
        servidor.quit()

        flash("E-mail enviado com sucesso.")

    except Exception as e:
        flash(f"Erro ao enviar e-mail: {str(e)}")

    return redirect("/dashboard")


# ==================================================
# EXPORTAR EXCEL
# ==================================================
@app.route("/exportar/excel")
def exportar_excel():
    if not login_obrigatorio():
        return redirect("/")

    from openpyxl import Workbook
    from openpyxl.styles import Font

    con = conectar()
    cur = con.cursor()

    cur.execute("""
        SELECT nome, cpf, telefone, email, status
        FROM clientes
        ORDER BY nome
    """)
    clientes = cur.fetchall()

    cur.close()
    con.close()

    wb = Workbook()
    ws = wb.active
    ws.title = "Clientes"

    headers = ["Nome", "CPF", "Telefone", "Email", "Status"]

    for i, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=i, value=h)
        c.font = Font(bold=True)

    for linha, cliente in enumerate(clientes, 2):
        for col, valor in enumerate(cliente, 1):
            ws.cell(row=linha, column=col, value=valor)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        as_attachment=True,
        download_name="clientes.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


# ==================================================
# LOGOUT
# ==================================================
@app.route("/logout")
def logout():
    session.clear()
    flash("Sessão encerrada.")
    return redirect("/")


if __name__ == "__main__":
    app.run(debug=True)